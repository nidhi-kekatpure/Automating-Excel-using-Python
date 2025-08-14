# go to projects->open in-> expoloer->copy paste the transaction.xlsx in the folder
import openpyxl as xl # as is aliance, so instead of openpyxl we can write xl in code
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx') #we have loaded excel in wb variable
sheet = wb['Sheet1'] # write the sheet1 with capital s as its the same in excel
# now to access rows and coloumns in excel:
cell = sheet['a1']#a1 are the coordinates of the sheet OR
cell = sheet.cell(1,1)
print(cell.value) #returns transtion_id as its a value oin the 1st row of 1st coloumn(a1 cell)

print(sheet.max_row) #give the max rows in the sheet which has value


for row in range(1,sheet.max_row + 1): #sheet.max_row+1 coz its exculsive,
#sheet.max_row gives value 4, so the loop above will run for 1,2,3 not 4 if we dont write +1

    print(row)

#now we need the 3rd coloumn of price in transaction.xlsx, but not the 1st row values
#as it has header price in it

for rows_in_cell in range(2, sheet.max_row+1):
    cell = sheet.cell(rows_in_cell,3) # row as we want to iterate all the rows in coloumn 3
    print(cell.value)

for rows_cell in range(2,sheet.max_row+1):
    cell = sheet.cell(rows_cell,3) # get each row value coloumn 3 exculeding header
    corrected_price = cell.value * 0.9
    corrected_price_coloumn = sheet.cell(rows_cell,4) #adding a coloumn 4
    corrected_price_coloumn.value = corrected_price #putting correct price in price coloumn value

values = Reference(sheet,
          min_row = 2, # min_row=2 as it removes rows of header in xl
          max_row = sheet.max_row, # these two sentences will select all the values in rows and coloumns mentioned(from row 2, till last row and all coloumn
          min_col = 4, max_col = 4) #we want values from 4th coloumn only

chart = BarChart() #creating instance of class BarChart() in object chart
chart.add_data(values)
sheet.add_chart(chart, 'e2') #add chart to sheet and at location e2 in excel

wb.save('transactions2.xlsx')

#go to automationexcel folder, right click, open in exploere, check the folder automationexcel
# and check for transtions2.xlsx

# Add charts from the coloumn 4 values




