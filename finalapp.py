# go to projects->open in-> explorer->copy paste the transaction.xlsx in the folder
import openpyxl as xl # as is aliance, so instead of openpyxl we can write xl in code
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename) #we have loaded excel in wb variable
    sheet = wb['Sheet1'] # write the sheet1 with capital s as its the same in excel

    #now we need the 3rd coloumn of price in transaction.xlsx, but not the 1st row values
    #as it has header price in it

    for rows_cell in range(2,sheet.max_row+1):
        cell = sheet.cell(rows_cell,3) # get each row value coloumn 3 exculding header
        corrected_price = cell.value * 0.9
        corrected_price_coloumn = sheet.cell(rows_cell,4) #adding a coloumn 4
        corrected_price_coloumn.value = corrected_price #putting correct price in price coloumn value

    # Add charts from the coloumn 4 values, write this after import:
    # from openpyxl.chart import BarChart, Reference

    values = Reference(sheet,
              min_row = 2, # min_row=2 as it removes rows of header in xl
              max_row = sheet.max_row, # these two sentences will select all the values in rows and coloumns mentioned(from row 2, till last row and all coloumn
              min_col = 4, max_col = 4) #we want values from 4th coloumn only

    chart = BarChart() #creating instance of class BarChart() in object chart
    chart.add_data(values)
    sheet.add_chart(chart, 'e2') #add chart to sheet and at location e2 in excel

    new_filename = "new_transactions.xlsx"
    wb.save(new_filename)

#go to automationexcel folder, right click, open in explorer, check the folder automationexcel
# and check for transcations2.xlsx


process_workbook('transactions.xlsx')



